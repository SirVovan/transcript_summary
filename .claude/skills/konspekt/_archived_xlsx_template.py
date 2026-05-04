"""
ЭТАЛОННЫЙ СКРИПТ ГЕНЕРАЦИИ КАРТЫ СМЫСЛОВ
=========================================
Использовать как основу при каждой генерации xlsx.
НЕ переписывать функции — они решают конкретные баги openpyxl.

ТРЕБОВАНИЯ К ЯЧЕЙКЕ «СУТЬ»:
1. Жирный текст — через CellRichText/TextBlock (НЕ Font(bold=True) на ячейку)
2. Абзацы — разделяются пустой строкой (\n\n), кроме элементов списка
3. Списки — символ • или цифра с точкой, каждый элемент на новой строке (\n)
4. Шрифт — Georgia 11pt (читабельнее для длинных текстов, на 1pt крупнее обычных ячеек)

ШРИФТЫ:
- Обычные ячейки (тайминг, спикер, тема, тип): Arial 10pt
- Ячейка «Суть»: Calibri Light 14pt
- Заголовки: Arial 10pt bold, заливка #D9E1F2
"""

import re
import zipfile
import os
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import xml.etree.ElementTree as ET

# Шрифты — задаём явно, не оставляем на дефолт Excel
FONT_PLAIN   = dict(name="Arial",   size=11)   # обычные ячейки
FONT_CONTENT = dict(name="Calibri Light", size=14)   # ячейка «Суть» — крупнее и читабельнее
FONT_HEADER  = dict(name="Arial",   size=11, bold=True)  # заголовки


def parse_bold(text):
    """
    Превращает текст с **маркерами** в CellRichText с жирными фрагментами.
    Шрифт Georgia 11pt применяется ко всем runs — обычным и жирным.
    \n в тексте сохраняются — patch_workbook() пропишет их как &#10; в XML.
    """
    parts = re.split(r'(\*\*.*?\*\*)', text, flags=re.DOTALL)
    blocks = []
    for part in parts:
        if not part:
            continue
        if part.startswith('**') and part.endswith('**'):
            inner = part[2:-2]
            if inner:
                # Жирный: Georgia 11pt bold
                blocks.append(TextBlock(
                    InlineFont(rFont="Calibri Light", sz=14, b=True),
                    inner
                ))
        else:
            # Обычный: Georgia 11pt
            blocks.append(TextBlock(
                InlineFont(rFont="Calibri Light", sz=14),
                part
            ))
    if not blocks:
        return CellRichText([TextBlock(InlineFont(rFont="Calibri Light", sz=14), '')])
    return CellRichText(blocks)


def calc_height(text, col_width=92, line_height=19, min_height=25):
    """Считает высоту строки по тексту ячейки «Суть»."""
    if not text:
        return min_height
    lines = text.split('\n')
    total = 0
    for line in lines:
        if line.strip() == '':
            total += 1
        else:
            total += max(1, math.ceil(len(line) / col_width))
    return max(total * line_height, min_height)


def patch_workbook(src_path, dst_path):
    """
    Патчит xlsx после сохранения:
    1. Исправляет абсолютные пути в workbook.xml.rels → относительные
    2. Восстанавливает переносы строк (\n → &#10;) в тексте ячеек
    3. Исправляет <b val="1"/> → <b/>, убирает пустые <rPr/>

    ВАЖНО: патчим XML как строку через str/re — НЕ через ET.parse/tostring.
    ET.tostring() повторно экранирует entities (&#1082; → &amp;#1082;)
    и кириллица выводится числами вместо букв.
    """
    with zipfile.ZipFile(src_path, 'r') as zin:
        files = {n: zin.read(n) for n in zin.namelist()}

    # Патч 1: относительные пути
    rels_key = 'xl/_rels/workbook.xml.rels'
    if rels_key in files:
        content = files[rels_key].decode('utf-8')
        content = content.replace('Target="/xl/worksheets/', 'Target="worksheets/')
        files[rels_key] = content.encode('utf-8')

    # Патч 2 + 3: sheet1.xml
    sheet_key = 'xl/worksheets/sheet1.xml'
    if sheet_key in files:
        xml = files[sheet_key].decode('utf-8')

        # <b val="1"/> → <b/>
        xml = xml.replace('<b val="1"/>', '<b/>')

        # Убираем пустые <rPr/>
        xml = xml.replace('<rPr/>', '')

        # Восстанавливаем переносы строк внутри <t>...</t>
        # Захватываем весь тег целиком — не добавляем лишний </t>
        def fix_t_newlines(m):
            attrs   = m.group(1)   # атрибуты тега (может быть пустым)
            content = m.group(2)   # текст внутри
            if '\n' not in content:
                return m.group(0)
            fixed = content.replace('\n', '&#10;')
            if 'xml:space' not in (attrs or ''):
                return '<t xml:space="preserve">' + fixed + '</t>'
            else:
                return '<t' + (attrs or '') + '>' + fixed + '</t>'

        xml = re.sub(
            r'<t( [^>]*)?>([^<]*)</t>',
            fix_t_newlines,
            xml,
            flags=re.DOTALL
        )

        files[sheet_key] = xml.encode('utf-8')

    with zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)


def verify_output(filepath):
    """
    Полная проверка качества файла перед выдачей пользователю.
    Файл НЕ выдавать если хотя бы одна проверка не прошла.
    """
    ok = True
    with zipfile.ZipFile(filepath, 'r') as z:
        sheet_xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
        rels_xml  = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')

    # 0. XML валиден
    try:
        ET.fromstring(sheet_xml)
        print("[OK] XML sheet1.xml: валиден")
    except ET.ParseError as e:
        print(f"[ОШИБКА] XML sheet1.xml повреждён: {e}")
        ok = False

    # 1. Жирный текст записался
    bold_count = len(re.findall(r'<b/>', sheet_xml))
    print(f"[OK] Жирных тегов: {bold_count}" if bold_count > 0
          else "[ОШИБКА] Жирное форматирование не записалось")
    if bold_count == 0:
        ok = False

    # 2. Переносы строк записались
    nl_count = sheet_xml.count('&#10;')
    print(f"[OK] Переносов строк: {nl_count}" if nl_count > 0
          else "[ПРЕДУПРЕЖДЕНИЕ] Переносы строк не найдены")

    # 3. Пути в rels относительные
    if '/xl/worksheets/' in rels_xml:
        print("[ОШИБКА] Абсолютные пути в workbook.xml.rels — Excel не откроет файл")
        ok = False
    else:
        print("[OK] Пути в rels: относительные")

    # 4. Нет двойных </t></t> — признак багованного патча
    double_t = len(re.findall(r'</t></t>', sheet_xml))
    if double_t > 0:
        print(f"[ОШИБКА] Найдено {double_t} двойных </t></t> — файл повреждён")
        ok = False
    else:
        print("[OK] Нет двойных </t></t>")

    # 5. Нет двойного экранирования entities (&amp;#)
    double_escape = len(re.findall(r'&amp;#\d+;', sheet_xml))
    if double_escape > 0:
        print(f"[ОШИБКА] Найдено {double_escape} двойных экранирований &amp;# — кириллица выведется числами")
        ok = False
    else:
        print("[OK] Двойного экранирования нет")

    # 6. Шрифт Georgia присутствует в ячейках Суть
    georgia_count = sheet_xml.count('Calibri Light')
    print(f"[OK] Шрифт Calibri Light: {georgia_count} вхождений" if georgia_count > 0
          else "[ПРЕДУПРЕЖДЕНИЕ] Шрифт Calibri Light не найден в XML")

    if ok:
        print(f"\n[ГОТОВО] Файл можно выдавать: {filepath}")
    else:
        print(f"\n[СТОП] Файл повреждён — НЕ выдавать до исправления")
    return ok


def generate_karta(segments, output_path):
    """
    Генерирует xlsx карты смыслов.

    segments — список словарей:
      num        (int)  — номер сегмента
      timing     (str)  — "00:00:00 - 00:03:39"
      speaker    (str)  — имя спикера
      topic      (str)  — тема
      block_type (str)  — тип блока
      content    (str)  — текст «Суть» с **жирным**, \n (строки), \n\n (абзацы)
      demo       (str)  — (опц.) «Демонстрации»

    ФОРМАТ content:
      Абзацы разделяются:    \n\n
      Элементы списка:       \n  (без пустых строк между ними)
      Жирный:                **текст**

    Пример:
      "Дмитрий объясняет **ключевую идею**.\n\n"
      "Три обязательных файла:\n"
      "• **project.md** — описание проекта\n"
      "• **CLAUDE.md** — правила для ИИ\n"
      "• **PLAN.md** — план задач"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Карта смыслов"

    has_demo = any(s.get('demo') for s in segments)

    headers = ["Сегмент #", "Тайминг", "Спикер(ы)", "Тема", "Тип блока", "Суть"]
    widths  = [4,            10,         22,           32,    20,           120]
    if has_demo:
        headers.append("Демонстрации")
        widths.append(45)

    header_font  = Font(**FONT_HEADER)
    header_fill  = PatternFill("solid", fgColor="FFD9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    plain_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    plain_font    = Font(**FONT_PLAIN)
    num_align     = Alignment(horizontal="center", vertical="top")

    for row_idx, seg in enumerate(segments, 2):
        content_text = seg.get('content', '')
        demo_text    = seg.get('demo', '')

        # Номер — по центру, Arial 10
        ws.cell(row=row_idx, column=1, value=seg['num']).alignment = num_align
        ws.cell(row=row_idx, column=1).font = plain_font

        # Простые колонки — Arial 10pt
        for col_idx, key in enumerate(['timing', 'speaker', 'topic', 'block_type'], 2):
            c = ws.cell(row=row_idx, column=col_idx, value=seg.get(key, ''))
            c.alignment = plain_align
            c.font = plain_font

        # «Суть» — Georgia 11pt, rich text с жирным
        ws.cell(row=row_idx, column=6).value     = parse_bold(content_text)
        ws.cell(row=row_idx, column=6).alignment = plain_align
        # Базовый шрифт ячейки тоже Georgia — на случай если rich text не применится
        ws.cell(row=row_idx, column=6).font = Font(name="Calibri Light", size=14)

        if has_demo:
            ws.cell(row=row_idx, column=7).value     = parse_bold(demo_text) if demo_text else ''
            ws.cell(row=row_idx, column=7).alignment = plain_align
            ws.cell(row=row_idx, column=7).font      = Font(**FONT_PLAIN)

        ws.row_dimensions[row_idx].height = calc_height(content_text)

    tmp_path = output_path + '.tmp.xlsx'
    wb.save(tmp_path)
    patch_workbook(tmp_path, output_path)
    os.remove(tmp_path)

    return verify_output(output_path)


# ─── ТЕСТ ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    test_segments = [
        {
            'num': 1,
            'timing': '00:00:00 - 00:03:39',
            'speaker': 'Дмитрий Ледовских',
            'topic': 'Зачем нужна документация',
            'block_type': 'Концепция',
            'content': (
                'Дмитрий открывает урок с **ключевого противоречия**: ИИ всё забывает.\n\n'
                'Каждый новый чат — как **новый сотрудник в первый день**: талантливый, '
                'но не знает ничего о вашем проекте.\n\n'
                'Три обязательных файла:\n'
                '• **project.md** — описание проекта\n'
                '• **CLAUDE.md** — правила для ИИ\n'
                '• **PLAN.md** — план задач'
            ),
            'demo': '🔤 00:02:15 — промпт создания project.md'
        },
        {
            'num': 2,
            'timing': '00:03:39 - 00:05:45',
            'speaker': 'Дмитрий Ледовских',
            'topic': 'Два типа документации',
            'block_type': 'Концепция',
            'content': (
                'Дмитрий разграничивает **два типа** документации:\n\n'
                '1. **Проектная** — архив знаний: что за проект, для кого, проблема\n'
                '2. **Техническая** — описание конкретной функции перед реализацией\n\n'
                'Метафора: проектная — «**что я хочу построить**», '
                'техническая — «**как построить конкретную комнату**».'
            ),
        },
    ]

    generate_karta(test_segments, '/mnt/user-data/outputs/test_karta.xlsx')
